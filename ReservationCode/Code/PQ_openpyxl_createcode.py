import random
import time
import openpyxl
import configparser

from Code.myabc_db import MyABC_db

sl = MyABC_db('../Config/db_config.conf', 'TESTDB')

def code_do(number, stripling_name, initial, num, specification, product_name, channel):
    # def code_do(filepath, num, initial, specification, channels):
    # 文件名称中的时间
    now = time.strftime('%Y-%m-%d')
    filename = number + stripling_name + '-' + '预约编码' + now + '.xlsx'
    # filepath = 'D:\\PythonProject\\ReservationCode\\Codefiles\\' + filename
    filepath = '../Codefiles/' + filename
    channels = channel + '-' + stripling_name + '-' + product_name

    # 创建Excel文件
    book = openpyxl.Workbook()
    # 创建工作表
    booksheet = book.active
    booksheet.title = '预约编码'
    # 写标题栏
    booksheet['A1'] = '渠道名称'
    booksheet['B1'] = '预约规格'
    booksheet['C1'] = '预约编码'

    # 写入内容
    # row = 2
    #
    # for name, age in name2Age.items():
    #     booksheet.cell(row, 1).value = name
    #     booksheet.cell(row, 2).value = age
    #     row += 1

    creattime = time.strftime('%Y-%m-%d %H:%M:%S')

    i = 2
    while True:
        result = str(random.randint(0, 9999)).zfill(4)
        code = initial + result
        re = sl.select_record_fetchone("select * from codenumber where code = '{}'".format(code))
        # re = sl.select_record_fetchone("select * from codenumber where code = 'CFCCBJ0500111'")
        print('re:', re)
        if re is None:
            booksheet.cell(i, 1, channels)
            booksheet.cell(i, 2, specification)
            booksheet.cell(i, 3, code)
            sl.insert_record(
                "insert into codenumber(code,createtime)values ('{}','{}')".format(code, creattime))
            print(result)
            i += 1
        else:
            print('预约编码重复，不进入数据库和excel表格中')
        if i > num + 1:
            break
    book.save(filepath)
    return filepath


def get_co_config():
    config = configparser.ConfigParser()
    config_path = '../Config/co_config.conf'
    config.read(config_path, encoding='utf-8')
    # sec=config.sections()
    # op = config.options("code_project")
    # config.items("code_project")
    number = config.get('code_project', 'number')
    stripling_name = config.get('code_project', 'stripling_name')
    initial = config.get('code_project', 'initial')
    num = config.getint('code_project', 'num')
    specification = config.get('code_project', 'specification')
    product_name = config.get('code_project', 'product_name')
    channel = config.get('code_project', 'channel')
    return number, stripling_name, initial, num, specification, product_name, channel